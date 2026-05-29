#!/usr/bin/env python3
"""
extract-huertos-urbanos.py — ALI-02 Huertos urbanos comunitarios de Canarias.

Tarea: soberanía alimentaria + tejido vecinal. Cada huerto urbano es un
pliegue donde lo agrícola y lo civil se solapan: producción local de
verdura, espacio intergeneracional, escuela de agroecología, centro
informal de barrio. Para OCRE · POLIS los pintamos como pequeños puntos
verde-marrón con icono pala — pista de que ahí hay tierra activa.

Estrategia de fuentes (investigada 2026-05-27):

  1. **Datos Abiertos LPGC** — DATASET OFICIAL en
     `http://datosabiertos.laspalmasgc.es/repositorio/ambiente/huertos.xlsx`.
     Devuelve los 9 huertos municipales con coordenadas, número de
     parcelas, dimensión total y nº de hortelanos. Es nuestra ÚNICA
     fuente municipal con datos abiertos estructurados.

  2. **OSM Geofabrik canary-islands** — `landuse=allotments` /
     `landuse=community_garden` / `leisure=garden + garden:type`.
     Sondeo: 203 polígonos crudos, pero muchos son subdivisiones de
     parcela individual (mapeo paralelo en el PBF) sin nombre propio.
     Filtramos: requerimos `name` no vacío, o área >= 800 m² aprox para
     descartar parcelitas. Devuelve un puñado de huertos con identidad
     ('Huerto Urbano Siete Palmas', 'La union de Gánigo', y otros).

  3. **Telde, SC Tenerife, La Laguna, Arrecife** — NO publican datos
     abiertos JSON/CSV de huertos. Solo noticias web y un PDF
     normativo de 2013 sin listado. Por tanto: catálogo `_seed: true`
     curado a mano con los huertos confirmados por prensa local y
     comunicados municipales (Jinámar Telde, Sobradillo SCT, Añaza SCT,
     San Bonito / Cho Canino / San Matías La Laguna, azoteas Arrecife
     como proyecto piloto, etc.). Coords aproximadas al centro del barrio.

  4. **Equipamientos institucionales** — La Granja Agrícola Experimental
     del Cabildo de Gran Canaria (Arucas) no es huerto vecinal estricto,
     pero es referente didáctico/agroecológico y forma parte del ecosistema
     — se incluye marcada como `gestion: institucional`.

Salida: `public/data/huertos-urbanos-canarias.geojson`

Schema de propiedades:
  - nombre
  - gestion: "municipal" | "comunitario" | "asociativo" | "institucional"
  - mun (resuelto por point-in-polygon contra barrios-canonical.json)
  - barrio (cuando se conoce explícitamente)
  - area_estimada_m2 (entero, o None)
  - parcelas_total
  - parcelas_disponibles
  - hortelanos
  - acceso_publico (bool)  ← ¿se puede pedir parcela como vecino?
  - fuente (corta: "Datos Abiertos LPGC" | "OSM" | "curado")
  - _seed (bool — true para los curados sin verificar in situ)
  - foto_url (cuando viene de LPGC abierto)
"""

from __future__ import annotations

import json
import os
import sys
import unicodedata
from collections import Counter

ROOT = "/Users/panch/KOINOS-iso"
PBF_CANDIDATES = [
    f"{ROOT}/GEOFABRIK/canary-islands-latest.osm.pbf",
]
LPGC_XLSX = f"{ROOT}/scripts/_cache/huertos_lpgc.xlsx"
OUT = f"{ROOT}/public/data/huertos-urbanos-canarias.geojson"
BARRIOS = f"{ROOT}/public/data/barrios-canonical.json"

# Bbox Canarias (lng_min, lat_min, lng_max, lat_max).
BBOX = (-18.3, 27.5, -13.3, 29.5)

GENERATED_AT = "2026-05-27"


# =============================================================================
# Curado manual — huertos confirmados por prensa/comunicados pero sin datos
# abiertos publicados. `_seed: true` para señalar que las coords son
# aproximadas al centro del barrio y deben refinarse cuando el municipio
# publique el dataset.
# =============================================================================

CURATED_SEEDS = [
    # ---- Telde (Gran Canaria) ----
    {
        "nombre": "Huerto Comunitario Agroecológico de Jinámar",
        "gestion": "comunitario",
        "mun": "Telde",
        "barrio": "Valle de Jinámar",
        "lon": -15.4209, "lat": 28.0381,
        "area_estimada_m2": 1500,
        "parcelas_total": None,
        "parcelas_disponibles": None,
        "acceso_publico": True,
        "fuente": "curado",
        "nota": "Operado por vecinos dentro del Plan Integral del Valle de Jinámar — agroecológico, sin fitosanitarios químicos."
    },
    {
        "nombre": "Huerto Urbano de San Juan (Telde)",
        "gestion": "municipal",
        "mun": "Telde",
        "barrio": "San Juan",
        "lon": -15.4188, "lat": 27.9959,
        "area_estimada_m2": None,
        "parcelas_total": None,
        "parcelas_disponibles": None,
        "acceso_publico": True,
        "fuente": "curado",
        "nota": "Anunciado por el área de Sector Primario del Ayuntamiento de Telde — pendiente de adjudicación."
    },

    # ---- Santa Cruz de Tenerife ----
    {
        "nombre": "Huerto Urbano El Sobradillo",
        "gestion": "asociativo",
        "mun": "Santa Cruz de Tenerife",
        "barrio": "El Sobradillo",
        "lon": -16.3015, "lat": 28.4324,
        "area_estimada_m2": 1200,
        "parcelas_total": None,
        "parcelas_disponibles": None,
        "acceso_publico": True,
        "fuente": "curado",
        "nota": "Proyecto Faro coordinado por la asociación Giro. Placa fotovoltaica, recogida de pluviales, accesos adaptados desde Av. Los Majuelos."
    },
    {
        "nombre": "Huerto Urbano Solidario de Añaza",
        "gestion": "asociativo",
        "mun": "Santa Cruz de Tenerife",
        "barrio": "Añaza",
        "lon": -16.2922, "lat": 28.4119,
        "area_estimada_m2": 900,
        "parcelas_total": None,
        "parcelas_disponibles": None,
        "acceso_publico": False,
        "fuente": "curado",
        "nota": "Fundación El Buen Samaritano. Prioriza familias con cargas familiares y colectivos vulnerables."
    },
    {
        "nombre": "Huerto de La Alegría de la Huerta",
        "gestion": "asociativo",
        "mun": "Santa Cruz de Tenerife",
        "barrio": "La Alegría",
        "lon": -16.2421, "lat": 28.4684,
        "area_estimada_m2": None,
        "parcelas_total": None,
        "parcelas_disponibles": None,
        "acceso_publico": True,
        "fuente": "curado",
        "nota": "Asociación Agroecológica y Cultural — en pleno casco urbano, ejemplo de huerto-barrio."
    },

    # ---- San Cristóbal de La Laguna ----
    {
        "nombre": "Huerto Urbano de San Bonito",
        "gestion": "municipal",
        "mun": "San Cristóbal de La Laguna",
        "barrio": "San Bonito",
        "lon": -16.3287, "lat": 28.4824,
        "area_estimada_m2": None,
        "parcelas_total": None,
        "parcelas_disponibles": 0,
        "acceso_publico": True,
        "fuente": "curado",
        "nota": "Lista de espera permanente. Sorteo entre empadronados mayores de 16 años con > 2 años de empadronamiento."
    },
    {
        "nombre": "Huerto Urbano de Cho Canino",
        "gestion": "municipal",
        "mun": "San Cristóbal de La Laguna",
        "barrio": "Cho Canino",
        "lon": -16.3531, "lat": 28.4906,
        "area_estimada_m2": None,
        "parcelas_total": None,
        "parcelas_disponibles": 0,
        "acceso_publico": True,
        "fuente": "curado",
        "nota": "Red Municipal de Huertos Comunitarios de La Laguna."
    },
    {
        "nombre": "Huerto Urbano de San Matías",
        "gestion": "municipal",
        "mun": "San Cristóbal de La Laguna",
        "barrio": "San Matías",
        "lon": -16.3119, "lat": 28.4744,
        "area_estimada_m2": None,
        "parcelas_total": None,
        "parcelas_disponibles": 0,
        "acceso_publico": True,
        "fuente": "curado",
        "nota": "Red Municipal de Huertos Comunitarios — relaciones intergeneracionales como objetivo explícito."
    },

    # ---- Arrecife (Lanzarote) ----
    {
        "nombre": "Huertos Urbanos en Azoteas (Arrecife)",
        "gestion": "municipal",
        "mun": "Arrecife",
        "barrio": "Casco",
        "lon": -13.5476, "lat": 28.9628,
        "area_estimada_m2": None,
        "parcelas_total": None,
        "parcelas_disponibles": None,
        "acceso_publico": True,
        "fuente": "curado",
        "nota": "Proyecto piloto en azoteas + vermicompostaje. Talleres rotativos por barrios, gestión Medio Ambiente Arrecife."
    },
    {
        "nombre": "Huertos Comunitarios Permaterra (Lanzarote)",
        "gestion": "asociativo",
        "mun": "Arrecife",
        "barrio": None,
        "lon": -13.5481, "lat": 28.9572,
        "area_estimada_m2": None,
        "parcelas_total": None,
        "parcelas_disponibles": None,
        "acceso_publico": True,
        "fuente": "curado",
        "nota": "Asociación de permacultura, subvencionada por Cabildo Lanzarote dentro del programa Reserva de la Biosfera."
    },

    # ---- Equipamiento institucional referente ----
    {
        "nombre": "Granja Agrícola Experimental del Cabildo",
        "gestion": "institucional",
        "mun": "Arucas",
        "barrio": "Cardones",
        "lon": -15.5070, "lat": 28.1419,
        "area_estimada_m2": 60000,
        "parcelas_total": None,
        "parcelas_disponibles": None,
        "acceso_publico": True,
        "fuente": "curado",
        "nota": "No es huerto vecinal estricto pero es el centro agroecológico de referencia: biólogos, ingenieros agrónomos, mercado agrícola quincenal y talleres abiertos."
    },
]


# =============================================================================
# Helpers
# =============================================================================

def _norm(s: str) -> str:
    if not s:
        return ""
    s = s.strip().lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return s


def _load_barrios():
    """(mun_name, bbox, ring_outer) para point-in-polygon de municipio."""
    try:
        with open(BARRIOS, "r") as f:
            d = json.load(f)
    except FileNotFoundError:
        return []
    out = []
    for b in d.get("barrios", {}).values():
        mun = b.get("mun_name") or ""
        bbox = b.get("bbox")
        geom = b.get("geometria") or {}
        coords = geom.get("coordinates")
        if not (mun and bbox and coords):
            continue
        rings = []
        if geom.get("type") == "Polygon":
            rings.append(coords[0])
        elif geom.get("type") == "MultiPolygon":
            for poly in coords:
                rings.append(poly[0])
        for r in rings:
            out.append((mun, bbox, r))
    return out


def _point_in_ring(lon, lat, ring) -> bool:
    inside = False
    n = len(ring)
    j = n - 1
    for i in range(n):
        xi, yi = ring[i][0], ring[i][1]
        xj, yj = ring[j][0], ring[j][1]
        intersect = ((yi > lat) != (yj > lat)) and \
                    (lon < (xj - xi) * (lat - yi) / (yj - yi + 1e-12) + xi)
        if intersect:
            inside = not inside
        j = i
    return inside


def municipio_for(lon, lat, barrios):
    for mun, bbox, ring in barrios:
        if not (bbox[0] <= lon <= bbox[2] and bbox[1] <= lat <= bbox[3]):
            continue
        if _point_in_ring(lon, lat, ring):
            return mun
    return None


# =============================================================================
# 1) Datos Abiertos LPGC — el bloque autoritativo del archipiélago.
# =============================================================================

def load_lpgc():
    try:
        import openpyxl
    except ImportError:
        print("WARN: openpyxl no instalado, saltando LPGC", file=sys.stderr)
        return []
    if not os.path.exists(LPGC_XLSX):
        print(f"WARN: {LPGC_XLSX} no encontrado, saltando LPGC", file=sys.stderr)
        return []
    wb = openpyxl.load_workbook(LPGC_XLSX, data_only=True)
    # Usamos la hoja "Huertos (2)" que incluye la columna IMAGEN.
    sh_name = "Huertos (2)" if "Huertos (2)" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sh_name]
    rows = []
    headers = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [c.strip().upper() if isinstance(c, str) else c for c in row]
            continue
        if not row or not row[0] or str(row[0]).strip().upper() == "TOTAL":
            continue
        d = dict(zip(headers, row))
        nombre = str(d.get("HUERTO", "")).strip()
        lat = d.get("LATITUD")
        lon = d.get("LONGITUD")
        if not (nombre and isinstance(lat, (int, float)) and isinstance(lon, (int, float))):
            continue
        area = d.get("DIMENSION TOTAL")
        n_parc = d.get("Nº PARCELAS")
        libres = d.get("Nº PARCELAS LIBRES")
        hortelanos = d.get("Nº HORTELANOS")
        imagen = d.get("IMAGEN")
        rows.append({
            "nombre": f"Huerto Urbano {nombre}",
            "gestion": "municipal",
            "mun": "Las Palmas de Gran Canaria",
            "barrio": nombre,
            "lon": float(lon),
            "lat": float(lat),
            "area_estimada_m2": int(area) if isinstance(area, (int, float)) else None,
            "parcelas_total": int(n_parc) if isinstance(n_parc, (int, float)) else None,
            "parcelas_disponibles": int(libres) if isinstance(libres, (int, float)) else None,
            "hortelanos": int(hortelanos) if isinstance(hortelanos, (int, float)) else None,
            "acceso_publico": True,
            "fuente": "Datos Abiertos LPGC",
            "foto_url": imagen if isinstance(imagen, str) else None,
            "_seed": False,
        })
    return rows


# =============================================================================
# 2) OSM PBF — huertos con identidad (nombre propio o área significativa).
# =============================================================================

def _polygon_area_m2(coords):
    """Aproximación rápida de área en m² usando la fórmula del shoelace
    con la latitud media como factor de proyección. Para polígonos pequeños
    (< ~10 km) el error es <2%, suficiente para filtrar parcelitas."""
    if len(coords) < 3:
        return 0.0
    lat_mean = sum(c[1] for c in coords) / len(coords)
    cos_lat = max(0.1, abs(__import__("math").cos(__import__("math").radians(lat_mean))))
    a = 0.0
    for i in range(len(coords)):
        x1, y1 = coords[i]
        x2, y2 = coords[(i + 1) % len(coords)]
        a += (x1 * y2 - x2 * y1)
    # convertir grados² a m² con la métrica local
    return abs(a) * 0.5 * (111320.0 ** 2) * cos_lat


def load_osm():
    try:
        import osmium
    except ImportError:
        print("WARN: osmium no instalado, saltando OSM", file=sys.stderr)
        return []
    pbf = next((p for p in PBF_CANDIDATES if os.path.exists(p)), None)
    if not pbf:
        print("WARN: PBF no encontrado, saltando OSM", file=sys.stderr)
        return []

    class H(osmium.SimpleHandler):
        def __init__(self):
            super().__init__()
            self.rows = []

        def _is_huerto(self, tags):
            lu = tags.get("landuse", "")
            if lu in ("allotments", "community_garden"):
                return True
            leis = tags.get("leisure", "")
            gtype = tags.get("garden:type", "")
            if leis == "garden" and gtype in ("community", "allotment"):
                return True
            return False

        def node(self, n):
            tags = {t.k: t.v for t in n.tags}
            if not self._is_huerto(tags):
                return
            try:
                lon, lat = n.location.lon, n.location.lat
            except Exception:
                return
            if not (BBOX[0] <= lon <= BBOX[2] and BBOX[1] <= lat <= BBOX[3]):
                return
            self.rows.append({
                "kind": "node",
                "osm_id": f"node/{n.id}",
                "lon": lon, "lat": lat,
                "tags": tags,
                "area_m2": None,
            })

        def way(self, w):
            tags = {t.k: t.v for t in w.tags}
            if not self._is_huerto(tags):
                return
            try:
                coords = [(nd.lon, nd.lat) for nd in w.nodes if nd.location.valid()]
            except Exception:
                return
            if not coords:
                return
            lon = sum(c[0] for c in coords) / len(coords)
            lat = sum(c[1] for c in coords) / len(coords)
            if not (BBOX[0] <= lon <= BBOX[2] and BBOX[1] <= lat <= BBOX[3]):
                return
            area = _polygon_area_m2(coords)
            self.rows.append({
                "kind": "way",
                "osm_id": f"way/{w.id}",
                "lon": lon, "lat": lat,
                "tags": tags,
                "area_m2": area,
            })

    h = H()
    h.apply_file(pbf, locations=True)

    # Filtra: queremos huertos identificables — name propio O área >= 800 m².
    keep = []
    for r in h.rows:
        name = (r["tags"].get("name") or r["tags"].get("name:es") or "").strip()
        area = r["area_m2"] or 0
        if name and len(name) >= 4:
            r["name"] = name
            keep.append(r)
        elif area >= 800:
            r["name"] = ""
            keep.append(r)

    # Dedup por proximidad (≈ 60 m). El PBF a veces tiene el polígono "padre"
    # y N polígonos "parcela hija" muy cerca. Nos quedamos con el de más área
    # (o con el que tenga nombre).
    keep.sort(key=lambda r: (-(1 if r["name"] else 0), -(r["area_m2"] or 0)))
    used = []
    final = []
    for r in keep:
        skip = False
        for u in used:
            dx = (r["lon"] - u[0]) * 111000 * 0.88  # ≈ cos(28°)
            dy = (r["lat"] - u[1]) * 111000
            if dx * dx + dy * dy < 60 * 60:
                skip = True
                break
        if skip:
            continue
        used.append((r["lon"], r["lat"]))
        final.append(r)

    # Convierte a nuestro schema.
    out = []
    for r in final:
        tags = r["tags"]
        lu = tags.get("landuse", "")
        gtype = tags.get("garden:type", "")
        if lu == "community_garden" or gtype == "community":
            gestion = "comunitario"
        else:
            gestion = "municipal"  # allotments suele ser cesión municipal
        nombre = r["name"] or f"Huerto urbano (OSM {r['osm_id']})"
        out.append({
            "nombre": nombre,
            "gestion": gestion,
            "mun": None,  # se resuelve después con barrios-canonical
            "barrio": None,
            "lon": round(r["lon"], 6),
            "lat": round(r["lat"], 6),
            "area_estimada_m2": int(r["area_m2"]) if r["area_m2"] else None,
            "parcelas_total": None,
            "parcelas_disponibles": None,
            "acceso_publico": True,
            "fuente": "OSM",
            "osm_id": r["osm_id"],
            "_seed": not bool(r["name"]),
        })
    return out


# =============================================================================
# Pipeline principal
# =============================================================================

def main():
    print(f"=== ALI-02 Huertos urbanos · {GENERATED_AT} ===", file=sys.stderr)

    rows = []

    # 1. LPGC
    lpgc = load_lpgc()
    print(f"LPGC abiertos: {len(lpgc)}", file=sys.stderr)
    rows.extend(lpgc)

    # 2. OSM
    osm = load_osm()
    print(f"OSM filtrados: {len(osm)}", file=sys.stderr)
    rows.extend(osm)

    # 3. Curado
    for c in CURATED_SEEDS:
        d = dict(c)
        d.setdefault("_seed", True)
        rows.append(d)
    print(f"Curado (_seed): {len(CURATED_SEEDS)}", file=sys.stderr)

    # Dedup global por proximidad — LPGC ya tiene Jinámar, OSM también lo
    # mapea, y el curado puede pisarlos. Prioriza LPGC > OSM nombrado > curado.
    priority = {"Datos Abiertos LPGC": 0, "OSM": 1, "curado": 2}
    rows.sort(key=lambda r: (priority.get(r.get("fuente", "curado"), 9), 0 if r.get("nombre") else 1))
    used = []
    final = []
    for r in rows:
        skip = False
        for u in used:
            dx = (r["lon"] - u[0]) * 111000 * 0.88
            dy = (r["lat"] - u[1]) * 111000
            if dx * dx + dy * dy < 120 * 120:
                skip = True
                break
        if skip:
            continue
        used.append((r["lon"], r["lat"]))
        final.append(r)

    # Resolver municipio para los OSM (sin mun) — point-in-polygon.
    barrios = _load_barrios()
    print(f"Barrios lookup: {len(barrios)} anillos", file=sys.stderr)
    for r in final:
        if not r.get("mun"):
            mun = municipio_for(r["lon"], r["lat"], barrios)
            if mun:
                r["mun"] = mun

    # GeoJSON features.
    features = []
    by_mun = Counter()
    by_gestion = Counter()
    seed_count = 0
    for r in final:
        props = {
            "nombre": r["nombre"],
            "gestion": r.get("gestion") or "municipal",
            "mun": r.get("mun") or "",
            "barrio": r.get("barrio"),
            "area_estimada_m2": r.get("area_estimada_m2"),
            "parcelas_total": r.get("parcelas_total"),
            "parcelas_disponibles": r.get("parcelas_disponibles"),
            "hortelanos": r.get("hortelanos"),
            "acceso_publico": bool(r.get("acceso_publico", True)),
            "fuente": r.get("fuente", "curado"),
            "_seed": bool(r.get("_seed", False)),
        }
        if r.get("nota"):
            props["nota"] = r["nota"]
        if r.get("foto_url"):
            props["foto_url"] = r["foto_url"]
        if r.get("osm_id"):
            props["osm_id"] = r["osm_id"]
        by_mun[props["mun"] or "(sin municipio)"] += 1
        by_gestion[props["gestion"]] += 1
        if props["_seed"]:
            seed_count += 1
        features.append({
            "type": "Feature",
            "geometry": {"type": "Point", "coordinates": [r["lon"], r["lat"]]},
            "properties": props,
        })

    fc = {
        "type": "FeatureCollection",
        "name": "huertos-urbanos-canarias",
        "generated_at": GENERATED_AT,
        "source": (
            "Datos Abiertos LPGC (huertos.xlsx, autoritativo) + OSM Geofabrik "
            "canary-islands (landuse=allotments / community_garden filtrado por "
            "nombre o área>=800m²) + catálogo curado con huertos confirmados "
            "por prensa/comunicados municipales para Telde, SC Tenerife, "
            "La Laguna y Arrecife."
        ),
        "count": len(features),
        "seeds": seed_count,
        "by_gestion": dict(by_gestion),
        "by_municipio": dict(by_mun.most_common()),
        "features": features,
    }
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w") as f:
        json.dump(fc, f, ensure_ascii=False, separators=(",", ":"))

    print(f"\nWROTE {OUT}", file=sys.stderr)
    print(f"  total huertos: {len(features)}", file=sys.stderr)
    print(f"  _seed (curados): {seed_count}", file=sys.stderr)
    print(f"  por gestión: {dict(by_gestion)}", file=sys.stderr)
    print(f"  por municipio: {dict(by_mun.most_common(15))}", file=sys.stderr)


if __name__ == "__main__":
    main()
